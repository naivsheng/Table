import numpy as np
import openpyxl as op
import os
import csv
import pandas as pd
import time
from datetime import datetime

class TableReader(object):
    def __init__(self):
        '''
        # 1.读取 excel/csv 表格内容
        # 2.根据订货信息更新表格
        # 3.在供货商文档增加“上次预定时间”列，用以记录并计算下次预定时间
        # 在订货点击确认按钮后自动调用Refresh函数，更新总览表记录订货操作的具体日期
        # TBC
        '''
        self.FL = ['']
        df = pd.read_excel('Filialen.xlsx',header = 0)
        self.FL.extend(df.iloc[:,0])
        self.FL = {x: self.FL[x] for x in range(len(self.FL))}
        self.FLd = {self.FL[x]:x-1 for x in range(len(self.FL))}
        self.LF = ['']
        df = pd.read_excel('Lieferant.xlsx',header = 0)
        self.LF.extend(df.iloc[:,0])
        self.LFd = {self.LF[x]: x for x in range(len(self.LF))}

        self.table = []
        # path = 'H:\\py\\test\\goasia'
        path = os.getcwd()
        os.chdir(path)
        
        
    def Reader(self,files,sheet=None):
        # 更改目标目录
        # 读取表格数据
        if not sheet:
            sheet = 0
        workbook = op.load_workbook(files)
        # sheet_name = workbook.get_sheet_names() # 获取工作表
        # worksheet = workbook.get_sheet_by_name(sheets[sheet])
        worksheet = workbook.worksheets[sheet]      # 默认获取第一张工作表
        # 获取行、列信息
        rows = worksheet.max_row
        columns = worksheet.max_column
        df = pd.read_excel(files, header=0,sheet_name=sheet,index_col=0) #打开Excel--.xls文件
        #df = pd.read_excel(files,header=0,index_col=0)
        return(df)

    def Writer(self,file_name=None,path=None): # TODO
        # 根据FL、LF 构建新Kontrolle表，并根据总览表写入应订货信息
        if not file_name:
            week = time.strftime("%W")
            file_name = 'KW' + week + '.xlsx'
        else: week = file_name[2:4]
        if path:
            file_list = os.listdir(path)
        else:file_list = os.listdir('.')
        if file_name not in file_list:
            # 读取分店列表
            workbook=op.Workbook()
            worksheet = workbook.active
            for i in range(len(self.LF)):
                worksheet.cell(1, i+1,self.LF[i])
            for i in range(len(self.FL)):
                worksheet.cell(i+1,1,self.FL[i])
            worksheet.cell(1,1,'ID')
            workbook.save(filename=file_name)
            # '''
        files = 'Lieferant.xlsx'
        df = pd.read_excel(files, header=0)
        for j in self.LF:
            l=[]
            if j == '': continue
            for i in self.FL:
                # int(df.at[LFd[j]-1,FL[i]]) 收货周
                if not i: continue
                a = df.at[self.LFd[j]-1,self.FL[i]]
                if a == '-': continue               # 用于标记：跳过无需订货项
                elif not a: a = 0
                else: a = int(a[2:4]) # 截取KW数据
                if (int(week)-a) >= int(df.at[self.LFd[j]-1,'订货周期']): # 应订计算仅需订货周数
                    l.append(i-1)
            a = []
            a.append(self.LFd[j])
            self.Updata(file_name,l,a,'s')
        
    def Refresh(self,FL,LF,woche=None):
            # 允许进行补录：默认woche传入值为空，若有传入值则为补录，写入传入值信息
            # today = datetime.now().strftime('%Y-%m-%d') # 获取今天日期
            today = datetime.now().strftime('%d-%m-%Y') # 获取今天日期
            if not woche: # 判断是否为补录信息
                woche = time.strftime("%W")
                a = 'KW%2s %s' % (woche,today)
            else: # 补录信息，无需写入具体日期
                a = 'KW%2s' % (woche)
            workbook = op.load_workbook('Lieferant.xlsx')
            worksheet = workbook.active
            for fl in FL:
                for lf in LF:
                    # worksheet.cell(row=fl+2,column=lf+1).value = a # 写入总览表
                    worksheet.cell(row=lf+1,column=fl+3).value = a
            workbook.save('Lieferant.xlsx')
            # TODO： 7-10天（工作日）预估到货时间，跨度到星期一的货记入前一周

    def Updata_to_LF(self,week):   
        # 更新总览表
        # 读取*周信息，判断‘k/s’是否在info中，更新df1、构建新总览表
        # 总览表序号标记： （Cohe TR，01KS）行0列2
        # 周表序号标记： （01KS，COHE TR）行0列1
        file = 'KW' + week + '.xlsx'
        df = self.Reader(file)  # 读取周信息
        files = 'Lieferant.xlsx'
        df1 = self.Reader(files) # 读取总览表
        dic={}
        for i in df._stat_axis.values.tolist():
            for col in range(1,df.shape[1]):
                info = str(df.iat[i,col])
                # print('行：{} ,列: {},信息为{}'.format(i,col,info))
                if 'b' in info:
                    df1.iat[col-1,i+2] = week   # 更新订货KW
        df1.to_excel(files,index=None)
        
    def Updata(self,files,FL,LF,flag,path=None):
        # FL 传入门店文件行号列表；LF传入供货商列表;flag 确认操作状态
        if path:
            os.chdir(path)
        workbook = op.load_workbook(files)
        worksheet = workbook.active
        for fl in FL:
            for lf in LF:
                a = worksheet.cell(row=fl+2, column=lf+1).value
                if not a: a = ''
                a = a + flag
                worksheet.cell(row=fl+2, column=lf+1).value = a
        # print(FL,LF)
        workbook.save(files)
    def Update_new_LF(self,LF):
        workbook = op.load_workbook('Lieferant.xlsx') 
        worksheet = workbook.active  
        for i in range(len(LF)-1):
            worksheet.cell(1,i+2,LF[i+1])
        workbook.save('Lieferant.xlsx')
    def Update_new_FL(self,FL):
        workbook = op.load_workbook('Filialen.xlsx') 
        worksheet = workbook.active  
        for i in range(len(FL)-1):
            worksheet.cell(i+2,1,FL[i+1])
        workbook.save('Filialen.xlsx')
    
    def Rechange(self,files,path=None):
        # 更改目标目录
        # 读取表格数据
        # 表格转置(不能处理合并单元格的部分，会报错)
        if path:
            os.chdir(path)
        workbook = op.load_workbook(files)
        worksheet = workbook.active
        rows = worksheet.max_row
        columns = worksheet.max_column
        #a = np.array(columns,rows)
        l = []
        new_file = files.replace('xlsx','.csv')
        for col in worksheet.columns:
            for cell in col:
                data = cell.value
                l.append(data)
            with open(new_file,'a',encoding='utf-8',newline='') as f:
                csv.writer(f).writerow(l)
            l = []
        
        '''
        # 以行、列方式读取表格数据
        for row in worksheet.rows:
            for cell in row:
                print(cell.value,end=" ")
            print()
            print()
        '''
        
    def run(self,files):
        # 更改目标目录
        # 
        pass

    def quchong(self):
        path = 'H:\\py\\test\\goasia\\crawer'
        os.chdir(path)
        file_list = os.listdir(path)
        for files in file_list:
            if '.xlsx' in files: 
                df = TableReader().Reader(files)
                '''
                IsDuplicated = df.duplicated(subset=['name','price']) # 列名
                data = df.drop_duplicates(subset=['name','price'])
                wb = load_workbook(files)
                writer = pd.ExcelWriter(files,engine='openpyxl')
                writer.book = wb
                data.to_excel(writer, sheet_name='去重')
                writer.save()
                print(files,'finish')
                '''
                df1 = df.drop_duplicates(keep='first')
                df1.to_excel(files)
                print(files,'finish')

if __name__ == '__main__':
    # path = 'H:\\py\\test\\goasia'
    #files = 'KW07.xlsx'
    # TableReader().Reader(files)
    TableReader().quchong()
    