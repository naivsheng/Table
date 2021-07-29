# -*- coding: UTF-8 -*-
# __author__ = 'Yingyu Wang'

from tkinter import *
import tkinter as tk
from tkinter.font import Font
from tkinter.ttk import *
from tkinter.messagebox import *
import os
import pandas as pd
from TableReader import TableReader
# import time
from datetime import datetime,date
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import scrolledtext
import calendar

class Application_ui(Frame):
    #这个类仅实现界面生成功能，具体事件处理代码在子类Application中。
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.master.title('订货计划辅助')
        self.master.geometry('1000x800')
        # path = 'H:\\py\\test\\goasia'
        path = os.getcwd()
        # os.chdir(path)
        # file_list = os.listdir(path)
        self.FL = ['']
        df = pd.read_excel('Filialen.xlsx',header = 0)
        self.FL.extend(df.iloc[:,0])
        self.F = df.iloc[:,1] # 储存门店全称
        self.FL = {x: self.FL[x] for x in range(len(self.FL))}
        self.FLd = {self.FL[x]:x-1 for x in range(len(self.FL))}
        self.LF = ['']
        df = pd.read_excel('Lieferant.xlsx',header = 0)
        self.LF.extend(df.iloc[:,0])
        self.LFd = {self.LF[x]: x for x in range(len(self.LF))}
        self.createWidgets()
    

    def createWidgets(self):
        '''
            # 创建主窗口：当周应订情况总览
            以供应商为单位建立查询键
            文件打开
            
        '''
        self.top = self.winfo_toplevel()    # 主窗口 
        self.style = Style()
        l0 = tk.Label(self.top,text='当前状态',width=8)
        l0.place(x=10,y=10)
        self.var = tk.StringVar()    # 将label标签的内容设置为字符类型，接收b1_clicked函数的传出内容, 显示当前操作状态
        l = tk.Label(self.top, textvariable=self.var, bg='white', fg='black', font=('Arial', 12), width=5, height=1)
        l.place(x=80,y=10)
        l1 = tk.Label(self.top,text='KW=',width=5,height=1)
        l1.place(x=140,y=10)
        self.week = tk.StringVar()
        dt = datetime.now()
        woche = dt.strftime("%W")
        self.week.set(woche)
        entry_week = tk.Entry(self.top,textvariable=self.week,width=5)
        entry_week.place(x=180,y=10)
        # 建立下拉选框，选择分店
        l2 = tk.Label(self.top,text='分店',width=5,height=1)
        l2.place(x=240,y=10)
        valueF = StringVar()
        self.cbxf = tk.ttk.Combobox(self.top, width = 8, height = 20, textvariable = valueF,state='readonly') #, postcommand = self.show_select)
        # self.cbxf["value"] = list(self.FL.values())
        self.L = list(self.FL.values())
        self.cbxf['value'] = self.L
        self.cbxf.place(x=280,y=10)
        l6 = tk.Label(self.top,text='供应商',width=6,height=1)
        l6.place(x=400,y=10)
        valueL = StringVar()
        self.cbxl = tk.ttk.Combobox(self.top,width = 10, height = 20, textvariable = valueL,state='readonly')
        self.cbxl["value"] = self.LF
        self.cbxl.place(x=450,y=10)
        
        self.TabStrip1 = Notebook(self.top)
        # 标签页1 订货信息
        self.TabStrip1.place(relx=0.02,rely=0.05,relwidth=0.95,relheight=0.95)
        self.TabStrip1__Tab1 = Frame(self.TabStrip1)
        # 点击 查看 按钮，显示 应订、未订
        b1 = tk.Button(self.TabStrip1__Tab1, text='查看', font=('Arial', 12), width=10, height=1, command=self.bestell_checked)
        b1.place(x=750,y=25)
        # 点击 确认 按钮，将 已订 部分存入xls
        b2 = tk.Button(self.TabStrip1__Tab1, text='确认', font=('Arial', 12), width=10, height=1, command=self.bestell_confirm)
        b2.place(x=750,y=65)
        # 记录文件打开时间，用以判断是否已有更改，提醒操作人刷新最新数据
        tl1 = tk.Label(self.TabStrip1__Tab1,text='打开文件时间',width=12,font=('Arial',13))
        tl1.place(x=750,y=300)
        self.b_open_time = tk.StringVar()    
        tl2 = tk.Label(self.TabStrip1__Tab1, textvariable=self.b_open_time,font=('Arial', 14), width=12, height=1)
        tl2.place(x=750,y=350)
        tl3 = tk.Label(self.TabStrip1__Tab1,text='当前文件时间',width=12,font=('Arial',13))
        tl3.place(x=750,y=400)
        self.b_file_time = tk.StringVar()    
        tl4 = tk.Label(self.TabStrip1__Tab1, textvariable=self.b_file_time,font=('Arial', 14), width=12, height=1)
        tl4.place(x=750,y=450)
        l3 = tk.Label(self.TabStrip1__Tab1,text='应订',width=5,height=1)
        l3.place(x=60,y=10)
        sb = Scrollbar(self.TabStrip1__Tab1) # 给列表增加滚动条，以防过多数据
        self.blist1 = Listbox(self.TabStrip1__Tab1,width=20,height=45,yscrollcommand=sb.set)
        self.blist1.place(x=10,y=30)
        sb.config(command=self.blist1.yview)
        l4 = tk.Label(self.TabStrip1__Tab1,text='未订',width=5,height=1)
        l4.place(x=300,y=10)
        self.blist2 = Listbox(self.TabStrip1__Tab1,width=20,height=45,selectmode = MULTIPLE)
        self.blist2.place(x = 250,y = 30)
        l5 = tk.Label(self.TabStrip1__Tab1,text='本次已订',width=8,height=1)
        l5.place(x=540,y=10)
        self.blist3 = Listbox(self.TabStrip1__Tab1,width=20,height=45,selectmode = MULTIPLE)
        self.blist3.place(x = 500,y = 30)
        b3 = tk.Button(self.TabStrip1__Tab1, text='订货', font=('Arial', 12), width=5, height=1, command=self.b_show_select)
        b3.place(x=420,y=280)
        b4 = tk.Button(self.TabStrip1__Tab1, text='取消', font=('Arial', 12), width=5, height=1, command=self.b_delect_select)
        b4.place(x=420,y=320)
        info = '应订 显示由之前的数据计算出的*周应订货物；已订 以多选方式显示并记录*周本次操作前未订货物'
        label = tk.Label(self.TabStrip1__Tab1,text = info, fg='green',font=('Arial',12),width=500)
        label.pack(side=BOTTOM)
        self.TabStrip1.add(self.TabStrip1__Tab1, text='总览')
        
        # 标签页2 当天信息
        self.TabStrip1__Tab2 = Frame(self.TabStrip1)
        # 点击 查看 按钮，显示 应订、未订
        b1 = tk.Button(self.TabStrip1__Tab2, text='查看', font=('Arial', 12), width=10, height=1, command=self.tag_checked)
        b1.place(x=750,y=25)
        # 点击 确认 按钮，将 已订 部分存入xls
        b2 = tk.Button(self.TabStrip1__Tab2, text='确认', font=('Arial', 12), width=10, height=1, command=self.tag_confirm)
        b2.place(x=750,y=65)
        # 记录文件打开时间，用以判断是否已有更改，提醒操作人刷新最新数据
        tl1 = tk.Label(self.TabStrip1__Tab2,text='打开文件时间',width=12,font=('Arial',13))
        tl1.place(x=750,y=300)
        self.t_open_time = tk.StringVar()    
        tl2 = tk.Label(self.TabStrip1__Tab2, textvariable=self.t_open_time,font=('Arial', 14), width=12, height=1)
        tl2.place(x=750,y=350)
        tl3 = tk.Label(self.TabStrip1__Tab2,text='当前文件时间',width=12,font=('Arial',13))
        tl3.place(x=750,y=400)
        self.t_file_time = tk.StringVar()    
        tl4 = tk.Label(self.TabStrip1__Tab2, textvariable=self.t_file_time,font=('Arial', 14), width=12, height=1)
        tl4.place(x=750,y=450)
        l3 = tk.Label(self.TabStrip1__Tab2,text='应订',width=5,height=1)
        l3.place(x=60,y=10)
        sb = Scrollbar(self.TabStrip1__Tab2) # 给列表增加滚动条，以防过多数据
        self.tlist1 = Listbox(self.TabStrip1__Tab2,width=20,height=45,yscrollcommand=sb.set)
        self.tlist1.place(x=10,y=30)
        sb.config(command=self.tlist1.yview)
        l4 = tk.Label(self.TabStrip1__Tab2,text='未订',width=5,height=1)
        l4.place(x=300,y=10)
        self.tlist2 = Listbox(self.TabStrip1__Tab2,width=20,height=45,selectmode = MULTIPLE)
        self.tlist2.place(x = 250,y = 30)
        l5 = tk.Label(self.TabStrip1__Tab2,text='本次已订',width=8,height=1)
        l5.place(x=540,y=10)
        self.tlist3 = Listbox(self.TabStrip1__Tab2,width=20,height=45,selectmode = MULTIPLE)
        self.tlist3.place(x = 500,y = 30)
        b3 = tk.Button(self.TabStrip1__Tab2, text='订货', font=('Arial', 12), width=5, height=1, command=self.t_show_select)
        b3.place(x=420,y=280)
        b4 = tk.Button(self.TabStrip1__Tab1, text='取消', font=('Arial', 12), width=5, height=1, command=self.t_delect_select)
        b4.place(x=420,y=320)
        info = '应订 显示由之前的数据计算出的*周应订货物；已订 以多选方式显示并记录*周本次操作前未订货物'
        label = tk.Label(self.TabStrip1__Tab2,text = info, fg='green',font=('Arial',12),width=500)
        label.pack(side=BOTTOM)
        self.TabStrip1.add(self.TabStrip1__Tab2, text='当天')
        
        self.TabStrip1__Tab3 = Frame(self.TabStrip1)
        # 点击 生成 按钮，生成到货计划表
        b1 = tk.Button(self.TabStrip1__Tab3, text='生成', font=('Arial', 12), width=10, height=1, command=self.Ankunft)
        b1.place(x=700,y=25)
        self.scr = scrolledtext.ScrolledText(self.TabStrip1__Tab3, width=85, height=35,font=("隶书",14),bg='whitesmoke')    # 加入滚动条以输出多行文本
        self.scr.place(x=30,y=60)
        b2 = tk.Button(self.TabStrip1__Tab3, text='说明', font=('Arial', 12), width=10, height=1, command=self.Info)
        b2.place(x=800,y=25)
        
        self.TabStrip1.add(self.TabStrip1__Tab3,text='到货表')
        

class Application(Application_ui):
    #这个类实现具体的事件处理回调函数。界面生成代码在Application_ui中。
    def tab_change(self,*args):
        # 切换标签页清空列表
        self.blist1.delete(0,END)
        self.blist2.delete(0,END)
        self.blist3.delete(0,END)
        self.klist1.delete(0,END)
        self.klist2.delete(0,END)
        self.klist3.delete(0,END)
        self.rlist1.delete(0,END)
        self.rlist2.delete(0,END)
        self.rlist3.delete(0,END)
        self.elist1.delete(0,END)
        self.elist2.delete(0,END)
        self.elist3.delete(0,END)
        self.wlist1.delete(0,END)
        self.wlist2.delete(0,END)
        self.wlist3.delete(0,END)
        
    def __init__(self, master=None):
        Application_ui.__init__(self, master)
    def get_FileModifyTime(self,filePath):
        # 获取文件更改时间
        t = os.path.getmtime(filePath)
        return t

    def checked(self):
        self.var.set('查看')
        self.scr.delete(1.0, END)
        s = ''
        dt = datetime.now()
        local_time = dt.strftime("%m-%d %H:%M:%S")
        s = '查看时间：' + local_time + '\n'
        woche = int(self.week.get())
        if woche < 10:
            woche = '0' + str(woche)
        else: woche = str(woche)
        file = 'KW' + woche + '.xlsx'
        self.GetINFO(file)
        if not self.cbxf.get() and not self.cbxl.get():
            # 输出总览表
            # TODO: 保证格式 打印 s = s + str(self.dframe)
            self.var.set('Error')
            tk.messagebox.showwarning('Error','请选择分店或供应商')
        elif self.cbxf.get() and self.cbxl.get():
            i = self.FLd[self.cbxf.get()]
            info = str(self.dframe.at[i,self.cbxl.get()]) 
            s = s + 'KW' + woche + ' ' + self.cbxf.get() + '  ' + self.cbxl.get() + '  详细信息:\n\n'
            if 'b' in info:
                s = s + '已订货 '
                if 'r' in info:
                    s = s + '已有发票 '
                else: s = s + '未收发票 '
                if 'k' in info:
                    s = s + '已收货 '
                    if 'z' in info: 
                        s = s + '有拍照确认 '
                    else: s = s + '无拍照确认 '
                    if 'y' in info: 
                        s = s + '点货后需投诉 '
                        if 'w' in info:
                            s = s + '已投诉 '
                        elif 'i' in info:
                            s = s + '不需要投诉'
                    else: s = s + '一切正常'
                else: s = s + '未收货'
            else: s = s + '未订货'
        elif self.cbxf.get():
            s = s + 'KW' + woche + ' ' + self.cbxf.get() + '  详细信息:\n\n'
            i = self.FLd[self.cbxf.get()]
            for col in self.dframe.columns.values.tolist():
                if col == 'ID':continue
                elif col == 'Unnamed: 0': continue
                # s = s + col + (10-len(col)) * 2 * ' ' +': '
                c = '%10s' % col
                s = s + c + ' :'
                info = str(self.dframe.at[i,col])
                if 'b' in info:
                    s = s + '已订货 '
                    if 'r' in info:
                        s = s + '已有发票 '
                    else: s = s + '未收发票'
                    if 'k' in info:
                        s = s + '已收货 '
                        if 'z' in info: 
                            s = s + '有拍照确认 '
                        else: s = s + '无拍照确认 '
                        if 'y' in info: 
                            s = s + '点货后需投诉 '
                            if 'w' in info:
                                s = s + '已投诉 '
                            elif 'i' in info:
                                s = s + '不需要投诉 '
                        else: s = s + '一切正常 '
                    else: s = s + '未收货 '
                    
                else: s = s + '未订货 '
                s = s + '\n'
        else:
            s = s + 'KW' + woche + ' ' + self.cbxl.get() + '  详细信息:\n\n'
            for i in range(self.dframe.shape[0]):
                info = str(self.dframe[self.cbxl.get()][i])
                s = s + self.dframe.at[i,'ID'] + (10-len(self.dframe.at[i,'ID'])) * ' ' + ': '
                if 'b' in info:
                    s = s + '已订货 '
                    if 'r' in info:
                        s = s + '已有发票 '
                    else: s = s + '未收发票 '
                    if 'k' in info:
                        s = s + '已收货 '
                        if 'y' in info: 
                            s = s + '点货后需投诉 '
                        if 'w' in info:
                            s = s + '已投诉 '
                        elif 'i' in info:
                            s = s + '不需要投诉'
                        else: s = s + '一切正常'
                    else: s = s + '未收货'
                else: s = s + '未订货 '
                s = s + '\n'
        self.scr.insert(END,s)
        # self.status.set(s)
    def Info(self):
        self.var.set('info')
        self.scr.delete(1.0, END)
        # files = 'INFO.txt'
        files = 'INFO_Bestell.txt'
        s = '现有功能及使用说明\n'
        with open(files,'r',encoding='utf-8') as f1:
            line = f1.readline()
            while line:
                s = s + line
                line = f1.readline()
        # self.status.set(s)
        self.scr.insert(END,s)
    def updata(self):
        week = datetime.now().isocalendar()[1]
        if week < 10:
            week = '0' + str(week)
        else: week = str(week)
        file = 'KW' + week + '.xlsx'
        
        today = datetime.now().weekday() # 周一为0
        now = int(time.time())
        FileTime = os.path.getmtime(file) # 时间戳
        '''
        file_time = time.mktime(time.strftime("%Y-%m-%d %H:%M",FileTime))
        if today > 3 and (now - FileTime > 5):
            TableReader().Updata_to_LF(week) # 更新总览表
        '''
        file = 'KW' + week + '.xlsx'
        
        # path = 'H:\\py\\test\\goasia'
        # file_list = os.listdir(path)
        file_list = os.listdir('.')
        if file not in file_list: 
            TableReader().Writer()

    def GetINFO(self,file): # 读取文件信息
        self.dframe = TableReader().Reader(file) # 分店索引从0开始
    
    def wocheconfirm(self): # 弹出警告窗口，询问是否继续
        b = True # 默认操作时间为当前周
        a = True
        woche = int(self.week.get())
        now = datetime.now().strftime("%W")
        if woche < int(now):
            b = False
            a = tk.messagebox.askquestion(title='Warning',message='预定时间非当前周，请确认是否仍要继续')
        if not a: # 用户点击取消，本次操作不保存
            return False,b
        if woche < 10: # 格式化
            woche = '0' + str(woche)
        else: woche = str(woche)
        return woche,b

    def bestell_checked(self):
        # 点击 查看 后，从供货商文档中获取列表，对比总览表中的数据，显示本周应订
        # 显示n周应到货、未到货信息
        self.var.set('总览')
        self.blist1.delete(0,END)
        self.blist2.delete(0,END)
        self.blist3.delete(0,END)
        local_time = datetime.now().strftime("%m-%d %H:%M:%S")
        # local_time = time.strftime("%m-%d %H:%M:%S", time.localtime())
        self.b_open_time.set(local_time)
        woche = int(self.week.get())
        if woche < 10:
            woche = '0' + str(woche)
        else: woche = str(woche)
        file = 'KW' + woche + '.xlsx'
        FileTime = self.get_FileModifyTime(file)
        self.b_file_time.set(datetime.now().strftime("%m-%d %H:%M:%S"))
        # self.b_file_time.set(time.strftime("%m-%d %H:%M:%S",FileTime))
        self.GetINFO(file)
        if not self.cbxf.get() and not self.cbxl.get():
            self.var.set('Error')
            tk.messagebox.showwarning('Error','请选择分店或供应商')
        elif self.cbxf.get() and self.cbxl.get():
            i = self.FLd[self.cbxf.get()]
            if 's' not in str(self.dframe.at[i,self.cbxl.get()]):
                message = 'KW' + woche + ' ' + self.cbxf.get() + '  ' + self.cbxl.get() + '无须订货'
                a = tk.messagebox.askquestion(title='Warning',message=message)
                if a:
                    self.blist2.insert(END,self.cbxl.get())
            else:
                self.blist2.insert(END,self.cbxl.get())
        elif self.cbxf.get():
            # 确认门店(行)，以供货商(列)形式查询
            i = self.FLd[self.cbxf.get()]
            for col in self.dframe.columns.values.tolist():
                if col == 'ID': continue
                elif col == 'Unnamed: 0': continue
                info = str(self.dframe.at[i,col])
                if 's' in info:
                    self.blist1.insert(END,col)
                if 'b' not in info:
                    self.blist2.insert(END,col)
        else:
            # 确认供货商，以门店查询
            for i in range(self.dframe.shape[0]):
                info = str(self.dframe[self.cbxl.get()][i])
                if 's' in info: # 判断是否应订
                    # print(self.dframe.at[i,'ID'])
                    self.blist1.insert(END,self.dframe.at[i,'ID'])
                if 'b' not in info:
                    self.blist2.insert(END,self.dframe.at[i,'ID'])
    def bestell_confirm(self): # 订货确认按钮
        a = self.blist3.size()
        woche,b = self.wocheconfirm()
        file = 'KW' + woche + '.xlsx'
        FL_updata = []
        LF_updata = []
        if not self.blist3:
            # 没有选择，直接返回
            return True
        if self.cbxf.get():
            FL_updata.append(self.FLd[self.cbxf.get()])
            {LF_updata.append(self.LFd[self.blist3.get(i)]) for i in range(self.blist3.size())}
        else:
            LF_updata.append(self.LFd[self.cbxl.get()])
            {FL_updata.append(self.FLd[self.blist3.get(i)]) for i in range(self.blist3.size())}
        TableReader().Updata(file,FL_updata,LF_updata,'b')
        # 更新总览表日期
        if b: TableReader().Refresh(FL_updata,LF_updata)  # 在总览表写入订货周数、日期
        else: # 补录信息
            TableReader().Refresh(FL_updata,LF_updata,woche)
    def b_show_select(self,*args):
        a = self.blist2.size()
        for i in range(a):
            if(self.blist2.select_includes(a-1-i)) == True: # 判断是否选中list中的数据
                # 删除未订表中的相应项，并将其加入已订表中
                self.blist3.insert(END,self.blist2.get(a-1-i))
                self.blist2.delete(a-1-i)
    def b_delect_select(self,*args):
        a = self.blist3.size()
        for i in range(a):
            if(self.blist3.select_includes(a-1-i)) == True: # 判断是否选中list中的数据
                # 删除已订表中的相应项，并将其加入未订表中
                self.blist2.insert(END,self.blist3.get(a-1-i))
                self.blist3.delete(a-1-i)
            
    def tag_checked(self):
        self.tlist1.delete(0,END)
        self.tlist2.delete(0,END)
        self.tlist3.delete(0,END)
        woche = int(self.week.get())
        local_time = datetime.now().strftime("%m-%d %H:%M:%S")
        #local_time = time.strftime("%m-%d %H:%M:%S", time.localtime())
        today = datetime.now().weekday() + 1
        self.var.set('周%s'%today)
        file = 'LF_datum.xlsx'
        dframe_t = TableReader().Reader(file)
        if woche < 10:
            woche = '0' + str(woche)
        else: woche = str(woche)
        file = 'KW' + woche + '.xlsx'
        self.t_open_time.set(local_time)
        FileTime = self.get_FileModifyTime(file)
        #self.t_file_time.set(time.strftime("%m-%d %H:%M:%S",FileTime))
        self.t_file_time.set(datetime.now().strftime("%m-%d %H:%M:%S"))
        self.GetINFO(file)
        if not self.cbxf.get() and not self.cbxl.get():
            self.var.set('Error')
            tk.messagebox.showwarning('Error','请选择分店或供应商')
        elif self.cbxf.get() and self.cbxl.get():
            i = self.FLd[self.cbxf.get()]
            if 's' not in str(self.dframe.at[i,self.cbxl.get()]):
                message = 'KW' + woche + ' ' + self.cbxf.get() + '  ' + self.cbxl.get() + '无需订货'
                tk.messagebox.askquestion(title='Warn',message=message)
            else:
                self.tlist2.insert(END,self.cbxl.get())
        elif self.cbxf.get():
            # 确认门店(行)，以供货商(列)形式查询
            i = self.FLd[self.cbxf.get()]
            for col in self.dframe.columns.values.tolist():
                if col == 'ID': continue
                elif col == 'Unnamed: 0': continue
                info = str(self.dframe.at[i,col])
                tag = dframe_t.at[i,col]
                if 's' in info and tag == today:
                    self.tlist1.insert(END,col)
                    if 'b' not in info and tag == today:
                        self.tlist2.insert(END,col)
        else:
            # 确认供货商，以门店查询
            for i in range(self.dframe.shape[0]):
                info = str(self.dframe[self.cbxl.get()][i])
                tag = dframe_t[self.cbxl.get()][i]
                if 's' in info and tag == today: # 判断是否应订
                    self.tlist1.insert(END,self.dframe.at[i,'ID'])
                    if 'b' not in info:
                        self.tlist2.insert(END,self.dframe.at[i,'ID'])
    def tag_confirm(self):
        a = self.tlist3.size()
        woche,b = self.wocheconfirm()
        file = 'KW' + woche + '.xlsx'
        FL_updata = []
        LF_updata = []
        if not self.tlist3:
            # 没有选择，直接返回
            return True
        if self.cbxf.get():
            FL_updata.append(self.FLd[self.cbxf.get()])
            {LF_updata.append(self.LFd[self.tlist3.get(i)]) for i in range(self.tlist3.size())}
        else:
            LF_updata.append(self.LFd[self.cbxl.get()])
            {FL_updata.append(self.FLd[self.tlist3.get(i)]) for i in range(self.tlist3.size())}
        TableReader().Updata(file,FL_updata,LF_updata,'b')
        # 更新总览表日期
        if b: TableReader().Refresh(FL_updata,LF_updata)  # 在总览表写入订货周数、日期
        else: # 补录信息
            TableReader().Refresh(FL_updata,LF_updata,woche)
    def t_show_select(self,*args):
        a = self.tlist2.size()
        for i in range(a):
            if(self.tlist2.select_includes(a-1-i)) == True: # 判断是否选中list中的数据
                # 删除未订表中的相应项，并将其加入已订表中
                self.tlist3.insert(END,self.tlist2.get(a-1-i))
                self.tlist2.delete(a-1-i)
    def t_delect_select(self,*args):
        a = self.tlist3.size()
        for i in range(a):
            if(self.tlist3.select_includes(a-1-i)) == True: # 判断是否选中list中的数据
                self.tlist2.insert(END,self.tlist3.get(a-1-i))
                self.tlist3.delete(a-1-i)

    def nextmontag(self,woche):
        # TODO 给定KW返回周一、下周一
        # 根据返回值进行到货时间判断
        delta = datetime.now().weekday()    # 周一为0，即可视为delta
        w = datetime.now().strftime("%W")
        y = int(datetime.now().strftime('%Y'))
        if int(woche) != w:
            # 非本周订货,即表单为过去补录,需返回给定周的周一、下周一信息
            day_jan_4th = date(y, 1, 4)
            first_week_start = day_jan_4th - timedelta(days=day_jan_4th.isoweekday()-1)
            # Monday = datetime.combine(first_week_start + timedelta(weeks=int(woche)-1), datetime.time()) #tzinfo=tzinfo)
            Monday = first_week_start + timedelta(weeks=int(woche)-1)
            nextMonday = Monday + timedelta(days=7)
        else:
            today = date.today()   # datetime.date 格式 YYYY-mm-dd
            Monday = today - timedelta(days = delta)
            nextMonday = today + timedelta(days=7-delta)
        return Monday,nextMonday

    def Ankunft(self):
        
        self.var.set('生成中')
        file = 'LF_ankunft.xlsx' # 读取到货时间表
        dframe_a = TableReader().Reader(file)
        # 读取总览表替代*周订货表，以计算某些隔周到货的供货商
        woche = int(self.week.get())
        nextwoche = woche + 1
        if nextwoche < 10:
            woche = '0' + str(woche)
            nextwoche = '0' + str(nextwoche)
        elif nextwoche == 10:
            woche = '0' + str(woche)
        else: 
            woche = str(woche)
            nextwoche = str(nextwoche)
        file = 'KW' + woche + '.xlsx'
        self.GetINFO(file)  # 获取订货信息
        wb = Workbook()
        wb = Workbook(write_only=True)
        files = 'KW' + woche + ' Ankunft.xlsx'
        wb.save(files)
        file = 'Lieferant.xlsx' 
        df = TableReader().Reader(file)        # 读取总览表
        montag,nextmonday = self.nextmontag(woche) # 返回本周、下周一的datetime.time
        # for i in range(1,dframe_a.shape[0]-1): # 遍历门店
        for i in range(0,dframe_a.shape[0]):
            L = {}
            #L[self.F[i-1]] = ['Mo','Di','Mi','Do','Fr','是否退货','成功退货']   # 表头：门店全称
            L[self.F[i]] = ['Mo','Di','Mi','Do','Fr','是否退货','成功退货']   # 表头：门店全称
            wb = load_workbook(files)
            writer = pd.ExcelWriter(files,engine='openpyxl')
            writer.book = wb
            for j in range(1,dframe_a.shape[0]+1):    # 遍历供货商
                L[self.LF[j]] = ['','','','','','','']
                # data = df.at[j-1,self.FL[i]] # 总览表订货时间信息 [供货商序号，门店缩写]
                # data = df.iat[j-1,i+1]  # 用序号替代
                data = df.iat[j-1,i+2]  # 用序号替代
                tag = ''
                struct_day = 0 # 初始化
                if data == '-': continue # 跳过无需订货
                wochen = int(data[2:4]) # 订货时间KW信息
                try: # 读取总览表具体订货日期信息
                    tag = data[5:]
                    struct_day = datetime.date(datetime.strptime(tag, "%d-%m-%Y")) # 时间戳
                except:pass
                # 对订货时间进行判断
                try:    # 读取'LF_ankunft.xlsx' 到货时间。 W: 固定某天到货；T: 订货后n天到货
                    zeitraum = str(dframe_a.at[i,self.LF[j]])[1:]   # 读取日期信息
                    flag = str(dframe_a.at[i,self.LF[j]])[0]        # 读取标记信息
                    info = int(zeitraum)
                except:     
                    flag = 'X'  # 缺省
                # 'W'则仅判断KW
                if flag == 'W' and int(woche) == int(wochen):
                    if info > 10:
                        L[self.LF[j]][info // 10 - 1] = 'X'
                        L[self.LF[j]][info % 10 - 1] = 'X'
                    else: L[self.LF[j]][info-1]= 'X'
                elif flag == 'X' and int(woche) == int(wochen): 
                    # 缺省默认下周到货
                    L[self.LF[j]][4] = '？'   # 不确定具体a到货时间，在星期五标记'？'
                elif flag == 'T': # 订货后n天到货
                    # 根据订货时间、到货周期、周一时间戳 计算到货时间是否落在下周
                    info = info + 1
                    if not struct_day: # 没有订货日期数据
                        continue
                    else: 
                        if info >= 7: info = info + 2   #对到货时间加2个时间单位(跳过周末)
                        else: # 判断时间戳+到货时间是否在下周范围内
                            if struct_day + timedelta(days=info) > montag + timedelta(days=4):
                                info = info + 2
                        lfdatum = struct_day + timedelta(days=info)
                        if lfdatum >= nextmonday and lfdatum <= (nextmonday + timedelta(days=5)): 
                            inf = lfdatum - nextmonday
                            inf = inf.days-1
                            L[self.LF[j]][inf] = 'X'
                        elif lfdatum >=(nextmonday + timedelta(days=6)) and lfdatum <= (nextmonday + timedelta(days=7)):
                            # 跳过周末，对落在周末的时间在周五标记'？'
                            L[self.LF[j]][4] = '?'
                    # TODO 跳过节假日 创建节假日列表？
                    
                
            df1 = pd.DataFrame(L)
            df1 = df1.T
            df1.to_excel(writer, sheet_name=self.FL[i+1])
            writer.save()
            
        # 更改表头： 删除多余信息，写入KW
        workbook = load_workbook(files)
        i = 1
        today = datetime.today()    # YYYY-mm-dd hh:MM:ss.ms
        oneday = timedelta(days = 1)
        m1 = calendar.MONDAY
        while today.weekday() != m1:
            today += oneday
        nextMonday = today.strftime('%d.%m.%Y')
        today = today + oneday * 6
        nextSunday = today.strftime('%d.%m.%Y')
        kw = 'KW ' + str(nextwoche) + ' : ' + nextMonday + ' bis ' + nextSunday
        while True:
            worksheet = workbook.worksheets[i]
            for j in range(1,8):
                worksheet.cell(1,j + 1).value = ''
            worksheet.cell(1,1).value = kw
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
            i += 1
            try: workbook.worksheets[i]
            except: break
        ws = workbook["Sheet"]
        workbook.remove(ws)
        workbook.save(files)
        self.var.set('完成')
        self.scr.delete(1.0, END)
        today = datetime.today()
        s = '当前操作时间为： ' + str(today) + ' \n' + 'KW' + str(nextwoche) + '到货计划表已生成\n文件名为： ' + files
        self.scr.insert(END,s)
        file = 'KW' + nextwoche + '.xlsx'
        file_list = os.listdir('.')
        if file not in file_list: 
            TableReader().Writer(file)
            s = '\n已生成%s周订货表%s\n'% (nextwoche,file)
            self.scr.insert(END,s)

if __name__ == "__main__":
    top = Tk()
    Application(top).mainloop()
    