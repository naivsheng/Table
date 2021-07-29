'''
# -*- coding: UTF-8 -*-
# __Author__: Yingyu Wang
# __date__: 
'''
import openpyxl as op
import pandas as pd

# 更改目标目录
# 读取表格数据
files = "KW29 Bestellung KW30 Lieferung Übersicht.xlsx"
workbook = op.load_workbook(files)
#sheet_name = workbook.get_sheet_names() # 获取工作表
worksheet = workbook.worksheets[0]      # 获取第一张工作表
# 获取行、列信息
rows = worksheet.max_row
columns = worksheet.max_column
# df = pd.read_excel(files, header=0) #打开Excel--.xls文件
df = pd.read_excel(files,header=0,index_col=0)
# print(df.loc['17 DD'][4]) # 对某周数据表进行检索，以店为单位确认信息
# df.loc['17 DD'][4] = 'test'
worksheet.cell(9,6).value = 'test' # 更改单元格数据不影响条件格式
workbook.save(files)
# print(df.loc['17 DD'][4])